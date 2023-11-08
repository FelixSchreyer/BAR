import math
import multiprocessing
import pickle

import numpy as np
import openpyxl
import pandas as pd
import joblib
from matplotlib import pyplot as plt
from sklearn.metrics import log_loss, f1_score, accuracy_score, recall_score, precision_score, confusion_matrix
from sklearn.utils import compute_class_weight

lock = multiprocessing.Lock()

def transfer_X(X, W):
    '''
    Takes Feature Vector and Weight Vector. Returns Vector of Features multiplied with their weight.
    :param X: Feature Vector
    :type X: pandas.DataFrame
    :param W: Weight Vector
    :type W: numpy.ndarray
    :return: Weighted Feature Vector
    :rtype: pandas.DataFrame
    '''
    X_hat = X * W
    return X_hat


def loss_funct(W, X, y, punishment):
    y_true = y
    sample_weights = {
        # 1: 1.66667,
        # 2: 3.5,
        # 3: 0.53846154
        1: 1.75,
        2: 2.333333,
        3: 0.5
    }
    sample_weights = compute_class_weight('balanced', classes=np.unique(y), y=y)

    X_hat = transfer_X(X, W)
    model = load_model()
    output = model.predict_proba(X_hat)
    output_translated = map_output(output)
    loss = log_loss(y, output_translated, labels=np.array([0, 1, 2]), sample_weight=[sample_weights[y-1] for y in y_true])

    class_of_interest = 1
    predictions = np.argmax(output_translated, axis=1)
    # loss_p = precision_score(y_true, predictions, labels = [class_of_interest], #labels=np.unique(predictions),  # labels=np.array([0, 1, 2]),
    #                         average='micro', zero_division=0)

    # y_diff = output_translated - y_true
    # y_diff = output_translated
    true_labels_array = y_true.to_numpy()
    predicted_probabilities = output_translated[np.arange(len(output_translated)), true_labels_array - 1]
    differences = np.where(predicted_probabilities > 0.6, 0, np.abs(0.6 - predicted_probabilities))
    differences_sum = differences.sum()

    predicted_classes = np.argmax(output_translated, axis=1)
    predicted_probabilities = output_translated[np.arange(len(output_translated)), true_labels_array - 1]
    predicted_classes = output_translated[np.arange(len(output_translated)), predicted_classes]
    differences = predicted_classes - predicted_probabilities

    # differences[differences < 0.1] = 0.1

    false_classifications = np.sum(y_true != predictions)

    differences_sum = differences.sum()

    return loss + punishment * differences_sum


def load_model():
    '''
    Loads Model of Random Forest
    :return: sklearn model
    :rtype: sklearn.model
    '''
    model = joblib.load('../Notebooks/classification_model.joblib')
    return model


def map_output(y):
    '''
    Maps output classes from source domain to classes from target domain.
    Maps second and third class of source to second class of target domain.
    See BAR paper for further explanations.
    :param y: predicted classes
    :type y: pandas.DataFrame
    :return: predicted classes mapped to target domain
    :rtype: pandas.DataFrame
    '''
    y[:, 2] += y[:, 3]
    y = np.delete(y, 3, axis=1)
    y_norm = y / np.sum(y, axis=1)[:, np.newaxis]
    return y_norm


def get_U(W):
    '''
    Generates a vector uniformly drawn at random from a unit euclidean sphere of the dimension of W.
    :param W: Weight Vector
    :return: random vector of dimension dim(W)
    :rtype: numpy.ndarray
    '''
    U_j = np.random.normal(size=W.shape)
    # U_j /= np.linalg.norm(U_j)

    return U_j


def update_W(W, alpha, gradient, iteration):
    '''
    Updates W either linearly or following an e-Function.
    :param W: Old Weight Vector
    :type W: pandas.DataFrame
    :param alpha: learning rate
    :type alpha: float
    :param gradient: estimated gradient at W
    :type gradient: np.ndarray
    :param iteration: Iteration in optimization process for update via e-Function
    :type iteration: int
    :return: new Weights
    :rtype: pandas.DataFrame
    '''

    # linear
    W = W + alpha * gradient

    # e-Function
    #print(iteration)
    #iteration+=1
    #decay = 0.02
    #W = W + math.exp(-decay * iteration) * alpha * gradient
    #print(np.max(math.exp(-decay * iteration) * alpha * gradient))
    return W


def rename_columns(target, source):
    '''
    Converts the column names aka feature names from those of the target domain to those of the source domain.
    If the number of columns in the target domain is smaller than in the source domain, columns with value 0 are added.
    If the target domain has more columns, a problem arises, the target domain columns need to be reduced then.
    :param target: Dataframe of Target Domain Data
    :type target: pandas.DataFrame
    :param source: Dataframe of Source Domain Data
    :type source: pandas.DataFrame
    :return: Dataframe that contains target data in format of source data
    :rtype: pandas.DataFrame
    '''
    column_names_target = target.columns.tolist()
    column_names_source = source.columns.tolist()

    extra_columns = list(set(column_names_target) - set(column_names_source))

    source.columns = column_names_target[:len(column_names_source)]

    for column_name in extra_columns:
        if column_name not in source.columns:
            source[column_name] = 0

    df_final = source[target.columns]
    return df_final


def print_progress_bar(iteration, total, length=50):
    progress = (iteration / total)
    arrow = '=' * int(length * progress)
    spaces = ' ' * (length - len(arrow))
    print(f'\r[{arrow}{spaces}] {int(progress * 100)}% \n', end='', flush=True)


def plot_loss(loss_hist):
    plt.bar(range(len(loss_hist)), loss_hist)

    # Add labels and a title
    plt.xlabel('Data Points')
    plt.ylabel('Values')
    plt.title('Bar Plot of Floats')

    # Show the plot
    plt.show()


def store_weights(file_path, weights, loss, W, W_0):
    # Save the list as a binary file
    with open(file_path, "wb") as file:
        pickle.dump(weights, file)

    np.save("weights", weights)

    print("List saved to", file_path)
    df = pd.DataFrame(W - W_0)
    df.to_csv('output.csv', index=False)


def calculate_f1_score(W, X, y):
    y_true = y
    sample_weights = {
        # 1: 1.66667,
        # 2: 3.5,
        # 3: 0.53846154
        1: 1.75,
        2: 2.333333,
        3: 0.5
    }
    X_hat = transfer_X(X, W)
    model = load_model()
    output = model.predict_proba(X_hat)
    output_translated = map_output(output)
    predictions = np.argmax(output_translated, axis=1)
    result = f1_score(y_true=y_true, y_pred=predictions + 1, average='macro')
    false_classifications = np.sum(y_true != predictions + 1)
    return result


def write_to_excel(settings, path):
    keys_to_exclude = ['model', 'X', 'y', 'X_s']
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet3']
    next_row = sheet.max_row + 1

    for setting in settings:
        filtered_params = {key: value for key, value in setting.items() if key not in keys_to_exclude}

        new_data = list(filtered_params.values())

        for col, value in enumerate(new_data, start=1):
            sheet.cell(row=next_row, column=col, value=value)

        next_row += 1

    wb.save(path)
